"""read_file must produce a document's text, or fail with a route the caller can take.

`_read_text_file` falls back to latin-1, which decodes any byte sequence without
raising, so a .xlsx or .pdf read by path used to return decoded binary that looked
like content. Extraction for PDF/DOCX/PPTX already existed but was reachable only
through the chat-attachment pipeline: the identical file sitting in the workspace
had no route at all. What genuinely cannot be read still fails fast, and the
diagnostic names a route that exists *for the caller asking* -- a worker has
run_command, the orchestrator does not.
"""
import asyncio
import datetime

import pytest

from opalatex import token_usage, tools


@pytest.fixture(autouse=True)
def clean_context():
    token_usage.set_usage_listener(None)
    token_usage.set_context_scope("reset")
    token_usage.set_context_scope("read-file-document-tests")
    yield
    token_usage.set_usage_listener(None)
    token_usage.set_context_scope("reset")


class _Session:
    def __init__(self, num_ctx):
        self.model_params = {"num_ctx": num_ctx}
        self.history = []


def _read_file(path):
    """Call read_file past the @opalatex_tool FunctionBlock/permission wrapper."""
    return asyncio.run(tools.read_file._func(path))


def _prepare(monkeypatch, target, num_ctx=128000, used=1000):
    monkeypatch.setattr(tools, "_PROJECT_SESSION", _Session(num_ctx))
    token_usage.record_context_tokens(used)
    monkeypatch.setattr(tools, "_resolve_path", lambda p: str(target))


def _make_xlsx(path, rows=None):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planejamento"
    for row in rows or [["Aula", "Data", "Conteudo"],
                        [1, datetime.datetime(2026, 8, 26), "Introducao"]]:
        ws.append(row)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Documents are extracted
# ---------------------------------------------------------------------------

def test_a_spreadsheet_is_extracted_instead_of_refused(tmp_path, monkeypatch):
    target = _make_xlsx(tmp_path / "plan.xlsx")
    _prepare(monkeypatch, target)

    out = _read_file(str(target))
    assert "Text extracted from 'plan.xlsx'" in out
    assert "Planejamento" in out
    assert "Introducao" in out
    # openpyxl hands back datetimes; a column of midnight timestamps is noise.
    assert "2026-08-26" in out
    assert "00:00:00" not in out


def test_a_multi_sheet_workbook_labels_every_sheet(tmp_path, monkeypatch):
    import openpyxl
    target = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "First"
    wb.active.append(["a", "b"])
    second = wb.create_sheet("Second")
    second.append(["c"])
    wb.save(target)
    _prepare(monkeypatch, target)

    out = _read_file(str(target))
    assert "# Sheet: First" in out
    assert "# Sheet: Second" in out


def test_a_real_pdf_in_the_workspace_is_extracted_by_path(tmp_path, monkeypatch):
    """Extraction existed only for chat uploads; a workspace PDF had no route."""
    pymupdf = pytest.importorskip("pymupdf")
    target = tmp_path / "calendario.pdf"
    doc = pymupdf.open()
    page = doc.new_page()
    page.insert_text((72, 100), "Calendario Academico 2026.2", fontsize=16)
    page.insert_text((72, 140), "Inicio das aulas: 2026-08-26", fontsize=11)
    doc.save(str(target))
    doc.close()
    _prepare(monkeypatch, target)

    out = _read_file(str(target))
    assert "Text extracted from 'calendario.pdf'" in out
    assert "Calendario Academico 2026.2" in out
    assert "2026-08-26" in out


def test_a_corrupt_document_reports_the_extraction_failure(tmp_path, monkeypatch):
    target = tmp_path / "broken.docx"
    target.write_bytes(b"PK\x03\x04 not really a docx")
    _prepare(monkeypatch, target)

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    message = str(excinfo.value)
    assert "Error extracting text" in message
    assert "Do not retry read_file" in message


def test_a_document_with_no_extractable_text_says_so(tmp_path, monkeypatch):
    """An image-only PDF extracts to nothing; that must not read as an empty file."""
    target = tmp_path / "scan.pdf"
    target.write_bytes(b"%PDF-fake")
    _prepare(monkeypatch, target)
    monkeypatch.setattr(
        "opalatex.attachments.extract_document_text_from_path",
        lambda path, max_chars=None: "   \n  ",
    )

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    assert "no extractable text" in str(excinfo.value)


def test_oversized_extracted_text_never_points_at_read_content_pos(tmp_path, monkeypatch):
    """Paging a .xlsx by line would read the raw bytes of a zip container."""
    target = _make_xlsx(tmp_path / "big.xlsx",
                        rows=[[f"row {i}", "x" * 60] for i in range(4000)])
    _prepare(monkeypatch, target, num_ctx=8000, used=7000)

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    message = str(excinfo.value)
    assert "does not fit the remaining context budget" in message
    assert "do not call read_content_pos on it" in message
    assert "command-line" in message


# ---------------------------------------------------------------------------
# What genuinely cannot be read fails fast
# ---------------------------------------------------------------------------

def test_an_image_is_routed_to_analyze_image(tmp_path, monkeypatch):
    target = tmp_path / "figure.png"
    target.write_bytes(b"\x89PNG\r\n\x1a\n\x00\x00\x00rest")
    _prepare(monkeypatch, target)

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    assert "analyze_image" in str(excinfo.value)


def test_an_unknown_extension_is_sniffed_for_nul_bytes(tmp_path, monkeypatch):
    target = tmp_path / "blob.dat"
    target.write_bytes(b"header\x00\x00payload")
    _prepare(monkeypatch, target)

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    assert "not a text file" in str(excinfo.value)


def test_utf16_text_is_not_mistaken_for_binary(tmp_path, monkeypatch):
    """UTF-16 is full of NUL bytes; _read_text_file already handles it via the BOM."""
    target = tmp_path / "notes.tex"
    target.write_bytes("\\section{Olá}\n".encode("utf-16"))
    _prepare(monkeypatch, target)

    assert "section" in _read_file(str(target))


def test_a_legacy_spreadsheet_without_a_reader_still_fails_fast(tmp_path, monkeypatch):
    target = tmp_path / "old.xls"
    target.write_bytes(b"\xd0\xcf\x11\xe0\x00\x00legacy OLE2 payload")
    _prepare(monkeypatch, target)

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    message = str(excinfo.value)
    assert "binary file" in message
    assert "no reader for .xls" in message
    # The orchestrator has no terminal tool, so "convert it yourself" is a dead end.
    assert "you cannot convert it yourself" in message
    assert "command-line" in message
    assert "plan mode" in message


def test_a_worker_is_told_to_convert_an_unreadable_file_itself(tmp_path, monkeypatch):
    """A skill worker does have run_command, so it gets the reachable advice."""
    target = tmp_path / "old.xls"
    target.write_bytes(b"\xd0\xcf\x11\xe0\x00\x00legacy")
    _prepare(monkeypatch, target)
    monkeypatch.setattr(tools, "_IN_SKILL_WORKER", True)

    with pytest.raises(ValueError) as excinfo:
        _read_file(str(target))
    message = str(excinfo.value)
    assert "You have run_command" in message
    assert "run_skill" not in message


def test_an_image_attachment_still_short_circuits_before_the_binary_check(monkeypatch):
    monkeypatch.setitem(tools._RECENT_FILE_ATTACHMENTS, "shot.png", {"type": "image"})
    with pytest.raises(ValueError) as excinfo:
        _read_file("shot.png")
    assert "image attachment" in str(excinfo.value)


def test_a_plain_text_file_is_unaffected(tmp_path, monkeypatch):
    target = tmp_path / "main.tex"
    target.write_text("\\documentclass{article}\n", encoding="utf-8")
    _prepare(monkeypatch, target)

    assert _read_file(str(target)) == "\\documentclass{article}\n"


# ---------------------------------------------------------------------------
# Model resolution
# ---------------------------------------------------------------------------
# get_agent_model(name, default) returns `default if default is not None else
# <agents.yaml default>`, so passing "" does not mean "no preference": an empty
# string is not None, so it beats the fallback chain and resolves to a model of
# "", which reaches litellm as "LLM Provider NOT provided". resolve_agent_model
# is the single answer both get_agent_llm_kwargs and any ephemeral agent use.

class _ModelSession:
    def __init__(self, model="", worker_model=""):
        self.model = model
        self.worker_model = worker_model
        self.model_params = {"num_ctx": 128000}
        self.worker_model_params = {}
        self.history = []


def test_the_worker_role_resolves_the_project_worker_model(monkeypatch):
    from opalatex import config
    monkeypatch.setattr(tools, "_PROJECT_SESSION",
                        _ModelSession(model="ollama/main", worker_model="ollama/worker"))
    assert config.resolve_agent_model("worker") == "ollama/worker"


def test_the_worker_role_falls_back_to_the_project_model(monkeypatch):
    from opalatex import config
    monkeypatch.setattr(tools, "_PROJECT_SESSION", _ModelSession(model="ollama/main"))
    assert config.resolve_agent_model("worker") == "ollama/main"


def test_with_no_session_the_configured_default_still_resolves(monkeypatch):
    from opalatex import config
    monkeypatch.setattr(tools, "_PROJECT_SESSION", None)
    assert config.resolve_agent_model("worker"), "must never resolve to an empty model"


def test_kwargs_and_model_are_resolved_from_the_same_answer(monkeypatch):
    """Credentials must belong to the model that will actually run."""
    from opalatex import config
    seen = []
    real = config.resolve_agent_model
    monkeypatch.setattr(config, "resolve_agent_model",
                        lambda name: seen.append(name) or real(name))
    config.get_agent_llm_kwargs("worker")
    assert seen == ["worker"]
